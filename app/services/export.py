from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.infrastructure.db.models import Request
from app.utils.timezone import format_moscow


class ExportService:
    """Выгрузка данных по заявкам в Excel (.xlsx)."""

    EXPORT_DIR = Path("exports")
    FILE_TEMPLATE = "requests_{start:%Y%m%d}_{end:%Y%m%d}.xlsx"

    @staticmethod
    async def export_requests(
        session: AsyncSession,
        *,
        start: datetime,
        end: datetime,
    ) -> Path:
        ExportService.EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        filename = ExportService.FILE_TEMPLATE.format(start=start, end=end)
        file_path = ExportService.EXPORT_DIR / filename

        requests = await ExportService._load_requests(session, start=start, end=end)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Заявки"

        headers = ExportService._headers()
        sheet.append(headers)
        ExportService._format_header(sheet[1])
        ExportService._append_requests(sheet, requests)
        ExportService._autofit_columns(sheet, len(headers))

        workbook.save(file_path)
        return file_path

    @staticmethod
    async def _load_requests(
        session: AsyncSession,
        *,
        start: datetime,
        end: datetime,
    ) -> list[Request]:
        stmt = (
            select(Request)
            .options(
                selectinload(Request.object),
                selectinload(Request.contract),
                selectinload(Request.defect_type),
                selectinload(Request.specialist),
                selectinload(Request.engineer),
                selectinload(Request.master),
                selectinload(Request.customer),
            )
            .where(Request.created_at.between(start, end))
            .order_by(Request.created_at)
        )
        result = await session.execute(stmt)
        return result.scalars().all()

    @staticmethod
    def _headers() -> list[str]:
        return [
            "Номер",
            "Заголовок",
            "Статус",
            "Дата создания",
            "Заказчик",
            "Объект",
            "Адрес",
            "Договор",
            "Тип дефекта",
            "Контактное лицо",
            "Телефон контакта",
            "Специалист",
            "Инженер",
            "Мастер",
            "Назначение осмотра",
            "Осмотр завершён",
            "Назначение мастера",
            "Старт работ",
            "Завершение работ",
            "Срок устранения (дата)",
            "Плановый бюджет",
            "Фактический бюджет",
            "Расхождение бюджета",
            "Плановые часы",
            "Фактические часы",
            "Ссылка на карточку",
            "Номер строки листа",
        ]

    @staticmethod
    def _format_header(header_row):
        bold = Font(bold=True)
        centered = Alignment(horizontal="center", vertical="center")
        for cell in header_row:
            cell.font = bold
            cell.alignment = centered

    @staticmethod
    def _append_requests(sheet, requests: Iterable[Request]) -> None:
        for req in requests:
            planned_budget = float(req.planned_budget or 0)
            actual_budget = float(req.actual_budget or 0)
            sheet.append(
                [
                    req.number,
                    req.title,
                    req.status.value,
                    ExportService._format_datetime(req.created_at),
                    req.customer.full_name if req.customer else "",
                    req.object.name if req.object else "",
                    ExportService._format_address(req),
                    req.contract.number if req.contract else "",
                    req.defect_type.name if req.defect_type else "",
                    req.contact_person,
                    req.contact_phone,
                    req.specialist.full_name if req.specialist else "",
                    req.engineer.full_name if req.engineer else "",
                    req.master.full_name if req.master else "",
                    ExportService._format_datetime(req.inspection_scheduled_at),
                    ExportService._format_datetime(req.inspection_completed_at),
                    ExportService._format_datetime(req.master_assigned_at),
                    ExportService._format_datetime(req.work_started_at),
                    ExportService._format_datetime(req.work_completed_at),
                    ExportService._format_datetime(req.due_at, "%d.%m.%Y"),
                    planned_budget,
                    actual_budget,
                    actual_budget - planned_budget,
                    float(req.planned_hours or 0),
                    float(req.actual_hours or 0),
                    req.sheet_url or "",
                    req.sheet_row_number or "",
                ]
            )

    @staticmethod
    def _format_datetime(dt: datetime | None, fmt: str = "%d.%m.%Y %H:%M") -> str:
        return format_moscow(dt, fmt) or ""

    @staticmethod
    def _autofit_columns(sheet, columns_count: int) -> None:
        min_width = 10
        max_width = 40
        for col_idx in range(1, columns_count + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in sheet[column_letter]:
                value = cell.value
                if value is None:
                    continue
                max_length = max(max_length, len(str(value)))
            width = min(max(max_length + 2, min_width), max_width)
            sheet.column_dimensions[column_letter].width = width

    @staticmethod
    def _format_address(req: Request) -> str:
        parts: list[str] = []
        if req.object and req.object.name:
            parts.append(req.object.name)
        if req.address:
            parts.append(req.address)
        if req.apartment:
            parts.append(f"кв. {req.apartment}")
        return ", ".join(parts)
